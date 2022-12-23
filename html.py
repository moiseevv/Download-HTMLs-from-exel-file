import requests
from openpyxl import load_workbook as lwb
COLUMN_KADASTR = 5          #  5 - я колонка с кадастровым номером для названия будущего файла
COLUMN_LINK_ON_HTML = 45    #  45 - я колонка с link ссылкой на html файл


wb = lwb("Кт.xlsx")
ws = wb.active
row_count = ws.max_row
print('Число строк в файле = ', row_count)


for i in range(2, row_count + 1):
    print("Строка", i)
    try:
        link_html = ws.cell(i, COLUMN_LINK_ON_HTML).hyperlink.display  # 45 строка это ссылка на html
        print("link_html ", link_html)
        kadastr_num_file = ws.cell(i, COLUMN_KADASTR).value
        kadastr_num_file = kadastr_num_file.replace(':', "_")
        print('Кадастровый номер ', kadastr_num_file)
        saved_xml_file = str(f'file_html\\{kadastr_num_file}.html')
        r = requests.get(link_html)
        html = r.content

        f = open(saved_xml_file, 'wb')
        f.write(html)
        f.close()
    except AttributeError:
		print("ОШИБКА AttributeError")
        pass
    except UnicodeEncodeError:
		print("ОШИБКА UnicodeEncodeError")
        pass
    except FileExistsError:
		print("ОШИБКА FileExistsError")
        pass

